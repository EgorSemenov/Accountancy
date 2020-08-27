from accountancy.models import File, Class, Group, Note
from openpyxl import load_workbook

import xlrd
from openpyxl.workbook import Workbook


def cvt_xls_to_xlsx(src_file_path, dst_file_path):  # переделывает xls файлы в xlsx файлы
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(dst_file_path)


def parse_class(file_id, i_row, ws):  # проходясь по exel файлу разбирает его, извлекая информацию о классах\
    # (строчки начинаются со слов "по классу")\ и вызывая функции по разбору остальных строк.
    class_id = get_num_fr_str(ws['A' + str(i_row)].value)
    import_class(class_id, ws['A' + str(i_row)].value)
    i_row += 1
    ind = True
    while not str(ws['A' + str(i_row)].value).startswith('ПО КЛАССУ'):
        if int(ws['A' + str(i_row)].value) < 100:
            ind = True
            i_row += 1
            continue
        else:
            if ind:
                parse_group(i_row, ws)
                ind = False
            parse_note(file_id, class_id, i_row, ws)
        i_row += 1
    if str(ws['A' + str(i_row + 1)].value).startswith('БАЛАНС'):
        return i_row + 1
    else:
        i_row += 1
        temp = parse_class(file_id, i_row, ws)
        return temp


def import_class(class_id, value):  # записывает информацию о классе в бд
    Class(class_id, value).save()


def parse_group(i_row, ws):  # разбирает строчку с группой(строчка которая начинается с двузначного числа)
    import_group(int(cell_val('A', i_row, ws)[0:2]))


def import_group(code):  # записывает информацию о группе в бд
    Group(Group.objects.count() + 1, code).save()


def parse_note(file_id, class_id, i_row,
               ws):  # разбирает строчку с записью(строчка которая начинается с четырехзначного числа)
    import_note(file_id, class_id, Group.objects.count(), int(cell_val('A', i_row, ws)[2:4]), get_row(i_row, ws))


def import_note(file_id, class_id, group_id, note_code, values):  # записывает информацию о записи в бд
    Note(Note.objects.count() + 1, note_code, values[0], values[1], values[2], values[3], file_id,
         class_id, group_id).save()


def get_row(i_row,
            ws):  # достает часть строки, значения в которой всегда совпадают по типу (с плав. точкой), под строкой\
    # понимается массив значений ячеек, находящихся в одной строке
    row = []
    for i in 'BCDE':
        row.append(float(cell_val(i, i_row, ws)))
    return row


def get_num_fr_str(s):  # достает число из строки, нужен, чтобы достать номер класса
    word_list = s.split()
    for word in word_list:
        if word.isnumeric():
            num = int(word)
            break
    return num


def cell_val(letter, number, ws):  # достает значение из exel ячейки
    return ws[letter + str(number)].value


def parse_file_bd(f, path, name):  # основной метод вызывающий остальные функции, и доходящий внутри файла до места, с которого нужно начинать парсинг
    file_id = import_file(name)
    if file_id == 0:
        return -1
    else:
        if 'xls' in name:
            cvt_xls_to_xlsx(path, path + 'x')
            path = path + 'x'
        wb = load_workbook(path)
        ws = wb.active
        i_row = 1
        while True:
            if str(ws['A' + str(i_row)].value).startswith('КЛАСС'):
                break
            i_row += 1
        final_row = parse_class(file_id, i_row, ws)
        print('Парсинг и загрузка в бд завершена, в файле было ', final_row, 'строк.')
        # export_file(1)
        return 0


def import_file(name):  # записывает данные о файле в бд или возвращает -1 в обработчик, если файл с таким именем уже есть
    if File.objects.filter(name=name).count() == 0:
        file_id = File.objects.count() + 1
        File(file_id, name).save()
        return file_id
    else:
        return 0
